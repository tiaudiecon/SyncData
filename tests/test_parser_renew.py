import io
from datetime import date, datetime
import openpyxl
import pytest
from app.services.parser_renew import ler_renew, RegistroRenew

HEADERS = ["Status", "Tipo de Nota", "Nome Original", "Novo Nome",
           "Fornecedor Emitente", "CNPJ do Emissor", "Nº NF / Série",
           "Data de Emissão", "Valor da NF"]


def _xlsx(*linhas):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for ln in linhas:
        ws.append(ln)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_le_e_normaliza():
    arq = _xlsx(
        ["OK", "PRODUTO", "x.pdf", "E_x.pdf", "VITORIA HOSPITALAR LTDA",
         "39.362.611/0001-15", "202069 / 7", datetime(2026, 5, 11), 1800],
    )
    itens = ler_renew(arq)
    assert len(itens) == 1
    it = itens[0]
    assert isinstance(it, RegistroRenew)
    assert it.numero == "202069 / 7"
    assert it.numero_norm == "202069"
    assert it.cnpj_emissor == "39362611000115"   # máscara removida
    assert it.fornecedor == "VITORIA HOSPITALAR LTDA"
    assert it.emissao == date(2026, 5, 11)
    assert it.valor == 1800.0        # `Valor da NF` = valor de face (bruto)


def test_coluna_obrigatoria_faltando_gera_erro_claro():
    headers_sem_valor = [h for h in HEADERS if h != "Valor da NF"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers_sem_valor)
    ws.append(["OK", "PRODUTO", "x.pdf", "E_x.pdf", "VITORIA HOSPITALAR LTDA",
               "39.362.611/0001-15", "202069 / 7", datetime(2026, 5, 11)])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    with pytest.raises(ValueError, match="Valor da NF"):
        ler_renew(buf)


def test_captura_novo_nome_como_arquivo_pdf():
    arq = _xlsx(
        ["OK", "PRODUTO", "x.pdf", "E_2026-05-11_NF202069.pdf", "VITORIA HOSPITALAR LTDA",
         "39.362.611/0001-15", "202069 / 7", datetime(2026, 5, 11), 1800],
    )
    it = ler_renew(arq)[0]
    assert it.arquivo_pdf == "E_2026-05-11_NF202069.pdf"


def test_sem_coluna_novo_nome_fica_vazio():
    headers = [h for h in HEADERS if h != "Novo Nome"]
    wb = openpyxl.Workbook(); ws = wb.active; ws.append(headers)
    ws.append(["OK", "PRODUTO", "x.pdf", "VITORIA HOSPITALAR LTDA",
               "39.362.611/0001-15", "202069 / 7", datetime(2026, 5, 11), 1800])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    it = ler_renew(buf)[0]
    assert it.arquivo_pdf == ""
