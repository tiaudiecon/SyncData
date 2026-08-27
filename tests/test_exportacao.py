import io
import openpyxl
from app.services.exportacao import gerar_xlsx


def _resumo():
    return {"cnpj": "04541288000162", "data_hora": "10/08/2026 10:55",
            "total_universo": 2, "valor_total": 250.0, "qt_gerenciadas": 1,
            "qt_ressalva": 0, "qt_falta_lancar": 1, "qt_falta_arquivar": 1,
            "qt_canceladas": 0}


def _itens():
    return [
        {"numero": "100", "nome_fornecedor": "FORNEC A", "data_emissao": "03/07/2026",
         "valor_bruto": 150.0, "valor_liquido": 150.0, "status_lancamento": "ok",
         "status_arquivo": "ok", "detalhe_lancamento": "", "detalhe_arquivo": "",
         "veredito": "gerenciada", "eh_gerenciada": True, "tem_erro": False},
        {"numero": "101", "nome_fornecedor": "FORNEC B", "data_emissao": "04/07/2026",
         "valor_bruto": 100.0, "valor_liquido": 100.0, "status_lancamento": "falta",
         "status_arquivo": "falta", "detalhe_lancamento": "", "detalhe_arquivo": "",
         "veredito": "pendente", "eh_gerenciada": False, "tem_erro": True},
        {"numero": "102", "nome_fornecedor": "FORNEC C", "data_emissao": "05/07/2026",
         "valor_bruto": 200.0, "valor_liquido": 190.0, "status_lancamento": "diverg",
         "status_arquivo": "ok", "detalhe_lancamento": "valor diverge", "detalhe_arquivo": "",
         "veredito": "ressalva", "eh_gerenciada": False, "tem_erro": True},
    ]


def _itens_ricos():
    base = _itens()
    for it in base:
        it.update({"sieg_bruto": it["valor_bruto"], "sieg_liquido": it["valor_liquido"],
                   "sieg_imp": 0.0, "sp_bruto": it["valor_bruto"],
                   "sp_liquido": it["valor_liquido"], "sp_imp": 0.0, "tem_desconto": False,
                   "impostos": {"sieg": {"iss":0,"inss":0,"ir":0,"csrf":0,"descontos":0,
                                "base_calculo":0,"aliquota":0,"iss_retido":False,"total":0},
                                "spdata": {"iss":0,"inss":0,"ir":0,"csrf":0,"total":0}}})
    return base


def test_aba_resumo_separada_e_total_sem_cabecalho():
    # item 5: 1ª aba "Resumo" (metadados) e aba "Total" só com a tabela de dados
    wb = openpyxl.load_workbook(io.BytesIO(gerar_xlsx(_resumo(), _itens_ricos())))
    assert wb.sheetnames[0] == "Resumo"
    assert "Total" in wb.sheetnames
    assert wb["Total"]["A1"].value == "Nº NF"     # Total começa direto na tabela
    res = [str(c.value) for row in wb["Resumo"].iter_rows() for c in row if c.value]
    assert "Cliente" in res and "CNPJ" in res and "Erros" in res


def test_export_tem_coluna_cnpj_em_todas_as_abas():
    # item 2: a coluna CNPJ aparece no relatório (aba Total e aba Impostos)
    itens = _itens_ricos()
    for it in itens:
        it["cnpj_fornecedor"] = "11.222.333/0001-99"
    wb = openpyxl.load_workbook(io.BytesIO(gerar_xlsx(_resumo(), itens)))
    res = [str(c.value) for row in wb["Total"].iter_rows() for c in row if c.value]
    assert "CNPJ" in res and "11.222.333/0001-99" in res
    imp = [str(c.value) for row in wb["Impostos"].iter_rows() for c in row if c.value]
    assert "CNPJ" in imp and "11.222.333/0001-99" in imp


def test_export_tem_aba_impostos_e_moeda_numerica():
    import io, openpyxl
    conteudo = gerar_xlsx(_resumo(), _itens_ricos())
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert "Impostos" in wb.sheetnames
    ws = wb["Total"]
    # acha uma célula de valor (Bruto Sieg) e confirma que é número com formato moeda
    achou = False
    for row in ws.iter_rows():
        for cel in row:
            if isinstance(cel.value, (int, float)) and "R$" in (cel.number_format or ""):
                achou = True
    assert achou


def test_abas_por_filtro():
    # item 1: abas Gerenciadas e Erros (Ressalva/Pendentes viraram Erros).
    conteudo = gerar_xlsx(_resumo(), _itens())   # 100 gerenc., 101 pendente, 102 ressalva
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    for aba in ("Total", "Gerenciadas", "Erros", "Impostos"):
        assert aba in wb.sheetnames, aba
    assert "Ressalva" not in wb.sheetnames and "Pendentes" not in wb.sheetnames
    # Erros junta pendente (101) + ressalva (102); Gerenciadas só a 100
    erros = [str(c.value) for c in wb["Erros"]["A"] if c.value is not None]
    assert "101" in erros and "102" in erros and "100" not in erros
    ger = [str(c.value) for c in wb["Gerenciadas"]["A"] if c.value is not None]
    assert "100" in ger and "101" not in ger


def test_confronto_inverso_em_aba_propria():
    # item 1: SP sem SIEG NÃO fica na aba Resultado — vai para aba própria.
    itens = _itens() + [{
        "numero": "999", "nome_fornecedor": "SO NO SP", "data_emissao": "05/07/2026",
        "valor_bruto": 0.0, "valor_liquido": 0.0, "status_lancamento": "",
        "status_arquivo": "", "detalhe_lancamento": "", "detalhe_arquivo": "",
        "veredito": "sp_sem_sieg", "sp_extra": True, "sp_bruto": 50.0, "sp_liquido": 50.0,
    }]
    wb = openpyxl.load_workbook(io.BytesIO(gerar_xlsx(_resumo(), itens)))
    assert "SP sem SIEG" in wb.sheetnames
    res = [str(c.value) for c in wb["Total"]["A"] if c.value is not None]
    assert "999" not in res                       # não polui a aba principal
    inv = [str(c.value) for c in wb["SP sem SIEG"]["A"] if c.value is not None]
    assert "999" in inv


def test_impostos_marca_recalculo():
    # item 2: a aba Impostos marca a linha divergente do recálculo.
    itens = _itens_ricos()
    itens[0]["pendencia_sieg"] = True
    itens[0]["pendencia_itens"] = [
        {"nome": "IRPJ", "codigo": "1708", "esperado": 45.0, "apurado": 0.0}]
    wb = openpyxl.load_workbook(io.BytesIO(gerar_xlsx(_resumo(), itens)))
    ws = wb["Impostos"]
    cab = [c.value for c in ws[1] if c.value]
    assert "Divergência (recálculo)" in cab
    texto = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "esperado R$ 45,00" in texto and "IRPJ 1708" in texto


def test_rel02_export_impostos_do_detalhamento():
    from app.services.exportacao import gerar_xlsx_impostos
    conteudo = gerar_xlsx_impostos(_itens_ricos())
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    assert wb.sheetnames == ["Impostos"]
    cab = [c.value for c in wb["Impostos"][1] if c.value]
    assert "IRPJ 1708 (Sieg)" in cab and "PIS/COFINS/CSLL 5952 (Sieg)" in cab


def test_texto_do_cliente_nao_vira_formula_viva():
    # openpyxl grava string começando com '=' como FÓRMULA; a planilha viaja
    # para a contabilidade, então o texto tem que sair como TEXTO.
    itens = _itens_ricos()
    itens[0]["nome_fornecedor"] = "=HYPERLINK(\"http://x\",\"clique\")"
    itens[1]["nome_fornecedor"] = "+1+1"
    itens[2]["detalhe_lancamento"] = "@SUM(A1:A9)"
    wb = openpyxl.load_workbook(io.BytesIO(gerar_xlsx(_resumo(), itens)))
    for aba in ("Total", "Impostos"):
        ws = wb[aba]
        for row in ws.iter_rows():
            for cel in row:
                if isinstance(cel.value, str):
                    assert cel.data_type != "f", (aba, cel.coordinate, cel.value)
                    assert not cel.value.startswith(("=", "+", "@")), cel.value
    total = [c.value for c in wb["Total"]["B"] if isinstance(c.value, str)]
    assert "'=HYPERLINK(\"http://x\",\"clique\")" in total
    assert "'+1+1" in total


def test_item_com_divergencia_e_exportado_na_aba_conciliacao():
    conteudo = gerar_xlsx(_resumo(), _itens())
    wb = openpyxl.load_workbook(io.BytesIO(conteudo))
    aba = wb["Total"]
    valores = [str(c.value) for c in aba["A"] if c.value is not None]
    assert "102" in valores
