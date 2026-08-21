import io
import openpyxl
from tests.test_conciliar import _sieg_xlsx, _spdata_txt
from tests._fakes import montar_conciliacao


def test_montar_itens_enriquecido(client):
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    from app.database import SessionLocal
    from app.models import Conciliacao
    from app.routers.resultado import montar_resumo_e_itens
    db = SessionLocal()
    conc = db.query(Conciliacao).order_by(Conciliacao.id.desc()).first()
    resumo, itens = montar_resumo_e_itens(conc)
    db.close()
    assert itens
    it = itens[0]
    for chave in ("sieg_bruto", "sieg_liquido", "sieg_imp", "sp_bruto", "impostos", "tem_desconto"):
        assert chave in it
    assert it["numero"].isdigit()


def test_item2_cnpj_do_spdata_no_detalhe(client):
    # item 2: o item expõe o CNPJ como veio do SP Data (p/ comparação com o SIEG)
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    from app.database import SessionLocal
    from app.models import Conciliacao
    from app.routers.resultado import montar_resumo_e_itens
    db = SessionLocal()
    conc = db.query(Conciliacao).order_by(Conciliacao.id.desc()).first()
    _, itens = montar_resumo_e_itens(conc)
    db.close()
    casados = [i for i in itens if i["fornecedor_sp"]]          # notas achadas no SP Data
    assert casados, "esperava ao menos uma nota casada com o SP Data"
    it = casados[0]
    raw = it["impostos"]["spdata"]["cnpj"]                       # gravado pela persistência
    assert raw
    formatado = it["cnpj_fornecedor_sp"]
    assert "/" in formatado                                      # veio formatado (XX.XXX.XXX/XXXX-XX)
    assert "".join(c for c in formatado if c.isdigit()) == raw


def test_iss_so_aparece_quando_retido(client):
    # o ISS só deve aparecer no detalhe quando for RETIDO; se não, nem a linha aparece
    import json
    from app.database import SessionLocal, engine, Base
    from app.models import Conciliacao, ConciliacaoItem
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    conc = Conciliacao(cnpj="11", competencia="2026-07"); db.add(conc); db.flush()

    def _imp(iss_retido):
        return json.dumps({
            "sieg": {"iss": 120.0, "inss": 0, "ir": 0, "csrf": 0, "descontos": 0,
                     "base_calculo": 1000.0, "aliquota": 0, "iss_retido": iss_retido,
                     "optante_sn": False, "total": 0.0},
            "spdata": {"iss": 0, "inss": 0, "ir": 0, "csrf": 0, "total": 0}})

    def _item(numero, cnpj, iss_retido):
        return ConciliacaoItem(
            conciliacao_id=conc.id, numero=numero, cnpj_fornecedor=cnpj,
            nome_fornecedor="FORN " + numero, data_emissao="03/07/2026",
            valor_bruto=1000.0, valor_liquido=1000.0, imp_sieg=0.0,
            impostos_json=_imp(iss_retido), status_lancamento="ok", status_arquivo="ok",
            veredito="gerenciada", cancelada=False, sp_valor_bruto=1000.0, sp_valor_liquido=1000.0)

    db.add(_item("1", "11", iss_retido=False))   # informativo -> some
    db.add(_item("2", "22", iss_retido=True))    # retido -> mostra
    db.commit(); cid = conc.id; db.close()

    html = client.get(f"/resultado/{cid}").text
    assert "informativo" not in html               # nunca mostra "ISS informativo"
    assert html.count(">ISS</td>") == 1            # só a nota com ISS retido tem a linha


def test_formatar_periodo_conferencia():
    from app.services.tempo import formatar_periodo, formatar_data_br
    assert formatar_data_br("2026-07-01") == "01/07/2026"
    assert formatar_data_br("") == "" and formatar_data_br(None) == ""
    assert formatar_periodo("2026-07-01", "2026-07-15") == "01/07/2026 a 15/07/2026"
    assert formatar_periodo("2026-07-01", "") == "01/07/2026"   # só início
    assert formatar_periodo(None, None) == ""


def test_resumo_expoe_periodo(client):
    # print3: o período informado na criação aparece no resumo (p/ mostrar no Resultado)
    from app.database import SessionLocal, engine, Base
    from app.models import Conciliacao
    from app.routers.resultado import montar_resumo_e_itens
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    conc = Conciliacao(cnpj="11", competencia="2026-07",
                       periodo_inicio="2026-07-01", periodo_fim="2026-07-20")
    db.add(conc); db.commit(); db.refresh(conc)
    resumo, _ = montar_resumo_e_itens(conc)
    db.close()
    assert resumo["periodo"] == "01/07/2026 a 20/07/2026"


def test_resultado_tem_busca_e_grupos(client):
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    st = montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    html = client.get(f"/resultado/{st['conciliacao_id']}").text
    assert 'id="busca"' in html
    assert "Exportar .xlsx" in html                # export da conciliação
    assert "Impostos .xlsx" not in html            # botão de impostos removido


def test_impostos_consolidado_no_resultado(client):
    # consolidação: a quebra de impostos vive no expand da nota, no Resultado.
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    st = montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    html = client.get(f"/resultado/{st['conciliacao_id']}").text
    assert "linha-det" in html and "det-tab" in html      # expand com a quebra
    assert "PIS/COFINS/CSLL" in html and "IRPJ" in html   # tributos no expand
    assert 'onclick="toggleDet' in html                   # clicar p/ expandir


def test_menu_sem_impostos(client):
    # a tela Impostos saiu do menu lateral (tudo no Resultado).
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    html = client.get("/").text
    assert 'href="/impostos"' not in html


def test_rotular_dv_formata_divergencia():
    from app.services.formatacao import rotular_dv
    assert rotular_dv("bruto R$ 5.029,00 ≠ R$ 4.719,71") == "Bruto: R$ 5.029,00 ≠ R$ 4.719,71"
    assert rotular_dv("impostos R$ 187,94 ≠ R$ 0,00") == "Impostos: R$ 187,94 ≠ R$ 0,00"


def test_export_xlsx_do_resultado_funciona(client):
    # "Exportar .xlsx" do Resultado gera um .xlsx válido (server-side ok).
    import io, openpyxl
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    st = montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    r = client.get(f"/resultado/{st['conciliacao_id']}/planilha.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "Resultado" in wb.sheetnames


def test_run_habilita_downloads_no_webview():
    # o WebView2 bloqueia downloads por padrão; run.py precisa liberar.
    import pathlib
    run = pathlib.Path(__file__).resolve().parent.parent / "run.py"
    assert 'ALLOW_DOWNLOADS' in run.read_text(encoding="utf-8")


def test_resultado_mostra_botao_pdf(client):
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    st = montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    html = client.get(f"/resultado/{st['conciliacao_id']}").text
    assert "/pdf/" in html
    assert "Abrir" in html


def test_item3_selos_presenca_spdata_e_sieg(client):
    # item 3: coluna Lançam. traz dois selos (SP Data e SIEG), verde=consta.
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    st = montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    html = client.get(f"/resultado/{st['conciliacao_id']}").text
    assert "selo-sis" in html and "SP Data" in html and "SIEG" in html
    from app.database import SessionLocal
    from app.models import Conciliacao
    from app.routers.resultado import montar_resumo_e_itens
    db = SessionLocal()
    conc = db.query(Conciliacao).order_by(Conciliacao.id.desc()).first()
    _, itens = montar_resumo_e_itens(conc)
    db.close()
    for chave in ("consta_spdata", "consta_sieg", "optante_sn"):
        assert chave in itens[0]


def test_item6_renomeia_pendencia_para_divergencia_impostos(client):
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    st = montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    html = client.get(f"/resultado/{st['conciliacao_id']}").text
    assert "Divergência Impostos" in html
    assert "Pendência SIEG" not in html and "SIEG?" not in html


def test_resultado_e_export(client):
    client.post("/setup", data={"cnpj": "04541288000162", "razao_social": "HSS"})
    st = montar_conciliacao("04541288000162", _spdata_txt(), _sieg_xlsx("04541288000162"))
    destino = f"/resultado/{st['conciliacao_id']}"
    assert client.get(destino).status_code == 200
    planilha = client.get(destino + "/planilha.xlsx")
    assert planilha.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(planilha.content))
    assert "Resultado" in wb.sheetnames and "Impostos" in wb.sheetnames
