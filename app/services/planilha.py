import unicodedata
import openpyxl


def norm_texto(txt) -> str:
    t = unicodedata.normalize("NFKD", str(txt or "")).encode("ascii", "ignore").decode()
    return " ".join(t.lower().split())


def mapa_cabecalho(headers) -> "dict[str, int]":
    return {norm_texto(h): i for i, h in enumerate(headers) if h is not None}


def indice(mapa, *rotulos):
    for rotulo in rotulos:
        i = mapa.get(norm_texto(rotulo))
        if i is not None:
            return i
    return None


def exigir_colunas(indices: "dict[str, int | None]", mensagem):
    """Confere que cada coluna obrigatória foi encontrada no cabeçalho.
    `indices` mapeia nome-da-coluna -> índice (ou None se não encontrada).
    Levanta ValueError com `mensagem(nome)` para a primeira coluna faltante."""
    for nome, i in indices.items():
        if i is None:
            raise ValueError(mensagem(nome))


def abrir_planilha(arquivo):
    """Abre a 1ª aba em modo read-only/data-only. `arquivo` = caminho ou
    file-like (BytesIO). Retorna (headers:list, linhas:iterador de tuplas).

    Lê todas as linhas e FECHA o workbook antes de retornar: em read-only o
    openpyxl mantém o .xlsx aberto (handle) enquanto o iterador vive, e no
    Windows isso travaria mover/renomear/apagar o arquivo depois. Estes
    relatórios têm poucos milhares de linhas, então materializar é barato."""
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        todas = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    headers = list(todas[0]) if todas else []
    return headers, iter(todas[1:])
