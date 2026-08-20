import io
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_NAVY = "FF1C2B5E"
_BRANCO = "FFFFFFFF"
_CREME = "FFF4F0E5"
_LINHA = "FFD9D4C6"
_TINTA = "FF15182B"
# Cores de status — UM verde, UM amarelo, UM vermelho, UM cinza (item 4).
_VERDE = "FF1A6E4A"
_AMARELO = "FFB0791E"
_VERMELHO = "FFA8331C"
_CINZA = "FF9CA0B5"

_FILL_OK = (PatternFill("solid", fgColor="FFE4F1EA"), Font(color=_VERDE, size=10))
_FILL_WARN = (PatternFill("solid", fgColor="FFF8EED7"), Font(color=_AMARELO, size=10))
_FILL_ALERT = (PatternFill("solid", fgColor="FFF6E0DA"), Font(color=_VERMELHO, size=10))
_FILL_NEUTRO = (None, Font(color=_CINZA, size=10))

_FONTE_CAB = Font(bold=True, color=_BRANCO, size=11)
_FILL_CAB = PatternFill("solid", fgColor=_NAVY)
_FILL_ZEBRA = PatternFill("solid", fgColor=_CREME)
_FONTE_DADO = Font(color=_TINTA, size=10)
_BORDA = Border(bottom=Side(style="thin", color=_LINHA))
_CENTRO = Alignment(vertical="center")
_MOEDA = 'R$ #,##0.00'
_TRACO = "—"

# veredito -> texto legível (mesma leitura da tela de Resultado)
_SITUACAO = {
    "gerenciada": "Gerenciada", "ressalva": "Ressalva", "pendente": "Faltou lançar",
    "cancelada": "Cancelada", "sp_sem_sieg": "SP Data sem SIEG",
    "sp_duplicada": "Duplicada no SP Data",
}
_ROTULO_ARQ = {"ok": "OK", "diverg": "Divergência", "falta": "Não encontrada"}

# Relatório espelha a tela: SN, presença SP Data × SIEG, arquivo, situação e a
# marcação do recálculo (qual tributo do SIEG diverge do base × alíquota).
CABECALHOS = ["Nº NF", "Fornecedor", "SN", "Emissão", "Lançam. (SP Data)",
              "Bruto (Sieg)", "Líquido (Sieg)", "Impostos (Sieg)",
              "Bruto (SPData)", "Líquido (SPData)", "Impostos (SPData)",
              "SP Data", "SIEG", "Arquivo", "Divergências", "Situação",
              "Divergência (recálculo)"]
_COLS_MOEDA = (6, 7, 8, 9, 10, 11)          # 1-based: as 6 colunas de valor
_COLS_STATUS = (12, 13, 14, 16)             # SP Data, SIEG, Arquivo, Situação
_COL_DIVERG = 15
_COL_RECALC_STD = 17                         # "Divergência (recálculo)" no layout padrão


def _cap_componente(parte):
    """'bruto R$ x ≠ R$ y' -> 'Bruto: R$ x ≠ R$ y' (mesma cara da tela)."""
    cabeca, _sep, resto = parte.partition(" ")
    return (cabeca.capitalize() + ": " + resto) if resto else parte.capitalize()


def _divergencias(it):
    """Texto das divergências igual à tela: por frente, um componente por linha."""
    linhas = []
    for titulo, det in (("Lançamento", it.get("detalhe_lancamento")),
                        ("Arquivo", it.get("detalhe_arquivo"))):
        if det:
            linhas.append(titulo + ":")
            linhas.extend(_cap_componente(p) for p in det.split("; "))
    return "\n".join(linhas)


def _estilo_por_texto(txt):
    """Fill/fonte da célula de status conforme o texto (verde/amarelo/vermelho/cinza)."""
    t = str(txt)
    if t in ("OK", "Gerenciada"):
        return _FILL_OK
    if t in ("Divergência", "Ressalva"):
        return _FILL_WARN
    if t in ("Faltou", "Não encontrada", "Faltou lançar",
             "SP Data sem SIEG", "Duplicada no SP Data"):
        return _FILL_ALERT
    if t in (_TRACO, "Cancelada"):
        return _FILL_NEUTRO
    return None


def _linha_item(it):
    spx = it.get("sp_extra")
    canc = it.get("cancelada")
    return [
        it["numero"], it["nome_fornecedor"],
        "Sim" if it.get("optante_sn") else _TRACO,                       # SN (item 3)
        _TRACO if spx else it["data_emissao"],
        it.get("sp_data_lancamento") or "",
        _TRACO if spx else it.get("sieg_bruto", 0.0),
        _TRACO if spx else it.get("sieg_liquido", 0.0),
        _TRACO if spx else it.get("sieg_imp", 0.0),
        it["sp_bruto"] if it.get("sp_bruto") is not None else _TRACO,
        it["sp_liquido"] if it.get("sp_liquido") is not None else _TRACO,
        it["sp_imp"] if it.get("sp_imp") is not None else _TRACO,
        _TRACO if canc else ("OK" if it.get("consta_spdata") else "Faltou"),   # SP Data
        "OK" if it.get("consta_sieg") else "Faltou",                          # SIEG
        _TRACO if (canc or spx) else _ROTULO_ARQ.get(it["status_arquivo"], it["status_arquivo"]),
        _divergencias(it),
        _SITUACAO.get(it["veredito"], it["veredito"].capitalize()),           # Situação
        _texto_recalc(it),                                                   # Divergência (recálculo)
    ]


def _largura(ws, cabecalhos, linhas):
    for i in range(1, len(cabecalhos) + 1):
        maior = max([len(str(cabecalhos[i - 1]))]
                    + [len(str(l[i - 1])) for l in linhas if i - 1 < len(l)] + [8])
        ws.column_dimensions[get_column_letter(i)].width = min(maior + 2, 55)


def _escrever_aba(ws, itens, com_totais=None):
    linha_atual = 1
    if com_totais:
        for rotulo, valor in com_totais:
            ws.cell(linha_atual, 1, rotulo).font = Font(bold=True, color=_NAVY, size=10)
            cel_v = ws.cell(linha_atual, 2, valor)
            cel_v.font = _FONTE_DADO
            if isinstance(valor, float):                 # ex.: Valor total (bruto)
                cel_v.number_format = _MOEDA
            linha_atual += 1
        linha_atual += 1  # linha em branco

    for c, texto in enumerate(CABECALHOS, start=1):
        cel = ws.cell(linha_atual, c, texto)
        cel.font = _FONTE_CAB
        cel.fill = _FILL_CAB
        cel.alignment = _CENTRO
    ws.freeze_panes = ws.cell(linha_atual + 1, 1)
    primeira_dado = linha_atual + 1

    linhas = [_linha_item(it) for it in itens]
    for offset, linha in enumerate(linhas):
        r = primeira_dado + offset
        for c, valor in enumerate(linha, start=1):
            cel = ws.cell(r, c, valor)
            cel.font = _FONTE_DADO
            cel.border = _BORDA
            if offset % 2 == 1:
                cel.fill = _FILL_ZEBRA
            if c in _COLS_MOEDA and isinstance(valor, (int, float)):
                cel.number_format = _MOEDA
            if c == _COL_DIVERG:
                cel.font = Font(color=_VERMELHO, size=9.5)     # UM vermelho (item 4)
                cel.alignment = Alignment(wrap_text=True, vertical="top")
            if c == _COL_RECALC_STD:                           # marca o tributo divergente
                cel.alignment = Alignment(wrap_text=True, vertical="top")
                if valor and valor != _TRACO:
                    cel.font = Font(color=_VERMELHO, size=9.5)
            if c in _COLS_STATUS:
                estilo = _estilo_por_texto(valor)
                if estilo:
                    fill, fonte = estilo
                    if fill:
                        cel.fill = fill
                    cel.font = fonte
    # item 4: linha de TOTAIS das colunas de valor
    if linhas:
        tot_r = primeira_dado + len(linhas)
        ws.cell(tot_r, 1, "TOTAIS").font = Font(bold=True, color=_NAVY, size=10)
        for c in _COLS_MOEDA:
            soma = round(sum(l[c - 1] for l in linhas if isinstance(l[c - 1], (int, float))), 2)
            cel = ws.cell(tot_r, c, soma)
            cel.font = Font(bold=True, color=_TINTA, size=10)
            cel.number_format = _MOEDA
    _largura(ws, CABECALHOS, [[""] * len(CABECALHOS)] + linhas)
    ws.column_dimensions[get_column_letter(_COL_DIVERG)].width = 46   # texto multi-linha
    ws.column_dimensions[get_column_letter(_COL_RECALC_STD)].width = 44
    if itens:   # filtro no cabeçalho p/ conferência
        ult = get_column_letter(len(CABECALHOS))
        ws.auto_filter.ref = f"A{linha_atual}:{ult}{primeira_dado + len(itens) - 1}"


# Colunas SIEG e SPData lado a lado + a marcação do recálculo (item 2).
CAB_IMP = ["Nº NF", "Fornecedor",
           "ISS (Sieg)", "ISS (SPData)", "INSS (Sieg)", "INSS (SPData)",
           "IRPJ 1708 (Sieg)", "IRPJ 1708 (SPData)",
           "PIS/COFINS/CSLL 5952 (Sieg)", "PIS/COFINS/CSLL 5952 (SPData)",
           "Descontos", "Base de cálculo", "Alíquota",
           "Total (Sieg)", "Total (SPData)", "Divergência (recálculo)"]
_COL_RECALC = 16                                    # coluna de marcação do recálculo
# colunas do SIEG cujo valor o recálculo confere (p/ pintar quando diverge)
_COL_SIEG_TRIB = {"1708": 7, "5952": 9}


def _rs(v):
    """float -> 'R$ 1.234,56' (texto)."""
    return ("R$ " + f"{(v or 0):,.2f}").replace(",", "X").replace(".", ",").replace("X", ".")


def _texto_recalc(it):
    """Marcação do recálculo: 'IRPJ 1708: esperado R$x ≠ SIEG R$y; ...' (item 2).
    Fornecedor em exceção (item 6) leva o prefixo 'EXCEÇÃO: <motivo>'."""
    itens = it.get("pendencia_itens") or []
    txt = "; ".join(
        f"{p['nome']} {p['codigo']}: esperado {_rs(p['esperado'])} ≠ SIEG {_rs(p['apurado'])}"
        for p in itens)
    if it.get("excecao"):
        obs = it.get("excecao_obs") or ""
        return ("EXCEÇÃO: " + obs + (" · " + txt if txt else "")).strip()
    return txt or "—"


def _aba_impostos(ws, itens):
    for c, t in enumerate(CAB_IMP, start=1):
        cel = ws.cell(1, c, t); cel.font = _FONTE_CAB; cel.fill = _FILL_CAB; cel.alignment = _CENTRO
    ws.freeze_panes = "A2"
    linhas_dados = []
    for off, it in enumerate(itens):
        s = (it.get("impostos") or {}).get("sieg") or {}
        p = (it.get("impostos") or {}).get("spdata") or {}
        vals = [it["numero"], it["nome_fornecedor"],
                s.get("iss", 0), p.get("iss"), s.get("inss", 0), p.get("inss"),
                s.get("ir", 0), p.get("ir"), s.get("csrf", 0), p.get("csrf"),
                s.get("descontos", 0), s.get("base_calculo", 0), s.get("aliquota", 0),
                s.get("total", 0), p.get("total"), _texto_recalc(it)]
        linhas_dados.append(vals)
        r = 2 + off
        pend = bool(it.get("pendencia_sieg"))
        cods_div = {str(x.get("codigo")) for x in (it.get("pendencia_itens") or [])}
        for c, v in enumerate(vals, start=1):
            cel = ws.cell(r, c, v); cel.font = _FONTE_DADO; cel.border = _BORDA
            if off % 2 == 1:
                cel.fill = _FILL_ZEBRA
            if 3 <= c <= 15 and c != 13 and isinstance(v, (int, float)):   # 13 = Alíquota
                cel.number_format = _MOEDA
        # item 2: MARCA a linha divergente do recálculo (célula de marcação + tributos)
        cel_rec = ws.cell(r, _COL_RECALC)
        cel_rec.alignment = Alignment(wrap_text=True, vertical="top")
        if pend:
            fill_a, fonte_a = _FILL_ALERT
            cel_rec.fill = fill_a; cel_rec.font = Font(color=_VERMELHO, size=9.5)
            for cod, col in _COL_SIEG_TRIB.items():
                if cod in cods_div:
                    ws.cell(r, col).fill = fill_a
                    ws.cell(r, col).font = fonte_a

    # linha de totais
    total_r = 2 + len(itens)
    ws.cell(total_r, 1, "TOTAL").font = Font(bold=True, color=_NAVY, size=10)
    somas = {}
    for it in itens:
        s = (it.get("impostos") or {}).get("sieg") or {}
        p = (it.get("impostos") or {}).get("spdata") or {}
        for c, v in ((3, s.get("iss", 0)), (4, p.get("iss") or 0), (5, s.get("inss", 0)),
                     (6, p.get("inss") or 0), (7, s.get("ir", 0)), (8, p.get("ir") or 0),
                     (9, s.get("csrf", 0)), (10, p.get("csrf") or 0),
                     (11, s.get("descontos", 0)), (12, s.get("base_calculo", 0)),
                     (14, s.get("total", 0)), (15, p.get("total") or 0)):
            somas[c] = round(somas.get(c, 0) + (v or 0), 2)
    for c, v in somas.items():
        cel = ws.cell(total_r, c, v)
        cel.font = Font(bold=True, color=_TINTA, size=10)
        cel.number_format = _MOEDA
    _largura(ws, CAB_IMP, [[""] * len(CAB_IMP)] + linhas_dados)   # largura pelos valores reais
    ws.column_dimensions[get_column_letter(_COL_RECALC)].width = 52   # texto do recálculo


def _principais(itens):
    return [i for i in itens if not i.get("cancelada") and not i.get("sp_extra")]


def gerar_xlsx_impostos(itens: list) -> bytes:
    """Planilha só com o Detalhamento de Impostos (quebra por tributo)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Impostos"
    _aba_impostos(ws, _principais(itens))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def gerar_xlsx(resumo: dict, itens: list) -> bytes:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Resultado"
    totais = [
        ("Cliente", resumo.get("razao_social") or resumo["cnpj"]),
        ("CNPJ", resumo["cnpj"]),
        ("Competência", resumo.get("competencia") or _TRACO),
        ("Gerado em", resumo["data_hora"]),
        ("Total de notas (Sieg)", resumo["total_universo"]),
        ("Valor total (bruto)", resumo["valor_total"]),
        ("Gerenciadas", resumo["qt_gerenciadas"]),
        ("Com ressalva", resumo["qt_ressalva"]),
        ("Faltou lançar", resumo["qt_falta_lancar"]),
        ("Faltou arquivar", resumo["qt_falta_arquivar"]),
        ("Canceladas (informativo)", resumo["qt_canceladas"]),
        ("SP Data sem SIEG", resumo.get("qt_sp_sem_sieg", 0)),
        ("Duplicadas no SP Data", resumo.get("qt_sp_duplicadas", 0)),
    ]
    prin = _principais(itens)
    # Aba "Resultado" = universo do Sieg (Todas) — SEM canceladas/confronto inverso,
    # que vão para abas próprias (item 1: separar por filtro, não misturar).
    _escrever_aba(ws1, prin, com_totais=totais)

    # Uma aba por filtro da tela (só as que têm notas).
    categorias = [
        ("Gerenciadas", [i for i in prin if i["veredito"] == "gerenciada"]),
        ("Ressalva", [i for i in prin if i["veredito"] == "ressalva"]),
        ("Pendentes", [i for i in prin if i["veredito"] == "pendente"]),
        ("Divergência Impostos", [i for i in prin if i.get("pendencia_sieg")]),
        ("Exceções", [i for i in prin if i.get("excecao")]),
        ("Canceladas", [i for i in itens if i.get("cancelada")]),
        ("SP sem SIEG", [i for i in itens if i["veredito"] == "sp_sem_sieg"]),
        ("Duplicadas SP", [i for i in itens if i["veredito"] == "sp_duplicada"]),
    ]
    for nome, subset in categorias:
        if subset:
            _escrever_aba(wb.create_sheet(nome), subset)

    _aba_impostos(wb.create_sheet("Impostos"), prin)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
